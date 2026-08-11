"""
Extract the 2026 UoM staff survey exports into tidy CSVs under data/.

Source workbooks live in "Staff Survey 2026/" (Qualtrics/Peakon-style dashboard
exports; each sheet opens with a ~9-row metadata block giving survey name,
response count, panel count, participation rate and report date).

Outputs:
  data/meta.csv          one row: response_count, panel_count, participation,
                          report_produced (shared across all four source files)
  data/themes.csv        theme, scope, score, population, positive, neutral, negative
  data/questions.csv     question_order, theme, question, score, population,
                          positive, neutral, negative, driver_strength, driver_correlation
  data/questions_5scale.csv
                          question_order, theme, question, score, population,
                          strongly_disagree, disagree, neutral, agree, strongly_agree,
                          driver_strength, driver_correlation
  data/theme_comparisons.csv
                          theme, benchmark, score, diff   (theme-level scores vs.
                          named benchmarks, e.g. prior UoM surveys, HEI sector)
  data/question_comparisons.csv
                          theme, question, benchmark, score, diff   (question-level
                          equivalent of theme_comparisons.csv)
  data/org_scores.csv    granularity, org_unit, n_responses, row_type, theme, question, score
  data/org_deltas.csv    granularity, org_unit, n_responses, row_type, theme, question,
                          score, delta_pp   (delta_pp = percentage points vs. Overall)
  data/comments.csv      question, engagement, category_1, category_2, category_3,
                          sentiment, comment
  data/comments_themed.csv
                          as comments.csv, plus a `themes` column: the set of
                          survey themes (pipe-separated) implied by the comment's
                          category tags, per sources/theme_category_map.csv.
                          A comment maps to 0, 1, or several themes; the mapping
                          is a user-editorial judgment call, not derived from IDs
                          shared with the quantitative data.

Comments are free-text and may be identifying when cross-referenced with small
org units — see CLAUDE.md before publishing any comment-level breakdowns. Note
comments have NO org-unit column at all (this export was taken with no
hierarchy filter applied) — theme linkage is the only correlation available;
comments cannot be broken out by Division/Department/Sub-Division.
"""
from pathlib import Path

import openpyxl
import pandas as pd

SRC = Path("Staff Survey 2026")
OUT = Path("data")
OUT.mkdir(exist_ok=True)

OVERALL_XLSX = SRC / "Main University of Manchester Survey 2026 - Survey Overall.xlsx"
DIV_DEPT_XLSX = SRC / "Main University of Manchester Survey 2026_Comp Div-Dept.xlsx"
SUB_DIV_XLSX = SRC / "Main University of Manchester Survey 2026_Comp Sub-Div.xlsx"
COMMENTS_XLSX = SRC / "Main University of Manchester Survey 2026_AllComments.xlsx"
THEME_CATEGORY_MAP_CSV = Path("sources") / "theme_category_map.csv"


def _rows(ws, min_row, max_row=None, max_col=None):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col, values_only=True):
        yield row


def _trim(row):
    row = list(row)
    while row and row[-1] is None:
        row.pop()
    return row


def extract_meta():
    """Shared metadata block (rows 1-9) present at the top of every sheet."""
    wb = openpyxl.load_workbook(OVERALL_XLSX, read_only=True, data_only=True)
    ws = wb["Details"]
    fields = {}
    for row in _rows(ws, 1, 9):
        if row[0] and row[1] is not None:
            key = row[0].rstrip(":").strip().lower().replace(" ", "_")
            fields[key] = row[1]
    wb.close()
    df = pd.DataFrame([fields])
    df.to_csv(OUT / "meta.csv", index=False)
    print(f"meta.csv: {list(fields.keys())}")


def extract_themes():
    """Theme-level rows (11-19) from Details and Details 5 Scale sheets."""
    wb = openpyxl.load_workbook(OVERALL_XLSX, read_only=True, data_only=True)

    records = []
    ws = wb["Details"]
    header = _trim(next(_rows(ws, 11, 11)))[2:]  # Percentage Score, Population, Positive, Neutral, Negative
    seen = set()
    for row in _rows(ws, 12, 19):
        theme = row[0]
        if theme is None or theme in seen:
            continue
        seen.add(theme)
        records.append({
            "theme": theme,
            "score": row[2],
            "population": row[3],
            "positive": row[4],
            "neutral": row[5],
            "negative": row[6],
        })
    wb.close()

    df = pd.DataFrame(records)
    df.to_csv(OUT / "themes.csv", index=False)
    print(f"themes.csv: {len(df)} rows")


def extract_questions():
    """Question-level detail (Details sheet, 7-point scale) from row 23 onward."""
    wb = openpyxl.load_workbook(OVERALL_XLSX, read_only=True, data_only=True)
    ws = wb["Details"]

    records = []
    for row in _rows(ws, 23):
        row = _trim(row)
        if not row or row[0] is None:
            continue
        order, theme, question, score, population, positive, neutral, negative = row[:8]
        driver_strength = row[8] if len(row) > 8 else None
        driver_correlation = row[9] if len(row) > 9 else None
        records.append({
            "question_order": order,
            "theme": theme,
            "question": question,
            "score": score,
            "population": population,
            "positive": positive,
            "neutral": neutral,
            "negative": negative,
            "driver_strength": driver_strength,
            "driver_correlation": driver_correlation,
        })
    wb.close()

    df = pd.DataFrame(records)
    df.to_csv(OUT / "questions.csv", index=False)
    print(f"questions.csv: {len(df)} rows")


def extract_questions_5scale():
    """Question-level detail (Details 5 Scale sheet) from row 23 onward."""
    wb = openpyxl.load_workbook(OVERALL_XLSX, read_only=True, data_only=True)
    ws = wb["Details 5 Scale"]

    records = []
    for row in _rows(ws, 23):
        row = _trim(row)
        if not row or row[0] is None:
            continue
        (order, theme, question, score, population,
         strongly_disagree, disagree, neutral, agree, strongly_agree) = row[:10]
        driver_strength = row[10] if len(row) > 10 else None
        driver_correlation = row[11] if len(row) > 11 else None
        records.append({
            "question_order": order,
            "theme": theme,
            "question": question,
            "score": score,
            "population": population,
            "strongly_disagree": strongly_disagree,
            "disagree": disagree,
            "neutral": neutral,
            "agree": agree,
            "strongly_agree": strongly_agree,
            "driver_strength": driver_strength,
            "driver_correlation": driver_correlation,
        })
    wb.close()

    df = pd.DataFrame(records)
    df.to_csv(OUT / "questions_5scale.csv", index=False)
    print(f"questions_5scale.csv: {len(df)} rows")


def _parse_comparisons_block(ws, header_row, data_start_row, data_end_row, has_question_col):
    """Shared parser for a Comparisons-sheet block (theme summary or question detail).

    Both blocks share the same column layout: [Theme, (Question,) Filtered
    Results, <Benchmark>, <Benchmark> (Diff), <Benchmark>, <Benchmark> (Diff), ...].
    Tidies to long form: one row per (theme[, question], benchmark).
    """
    header = _trim(next(_rows(ws, header_row, header_row)))
    score_start = 3 if has_question_col else 2
    benchmark_cols = []  # (score_col_idx, diff_col_idx, benchmark_name)
    i = score_start + 1  # skip the "Filtered Results" column, handled separately below
    while i < len(header):
        benchmark_cols.append((i, i + 1 if i + 1 < len(header) else None, header[i]))
        i += 2

    records = []
    for row in _rows(ws, data_start_row, data_end_row):
        row = list(row)
        if not row or row[0] is None:
            continue
        theme = row[0]
        question = row[1] if has_question_col else None
        filtered_score = row[score_start - 1]
        records.append({
            "theme": theme, "question": question,
            "benchmark": "Filtered Results", "score": filtered_score, "diff": None,
        })
        for score_i, diff_i, name in benchmark_cols:
            score = row[score_i] if score_i < len(row) else None
            diff = row[diff_i] if diff_i is not None and diff_i < len(row) else None
            if score == "n/a":
                score = None
            if diff == "n/a":
                diff = None
            records.append({
                "theme": theme, "question": question,
                "benchmark": name, "score": score, "diff": diff,
            })
    return pd.DataFrame(records)


def extract_comparisons():
    """Comparisons sheet: theme- and question-level scores against named benchmarks.

    The sheet has two stacked tables sharing the same benchmark columns:
    rows 11-19 (theme summary, one row per theme) and rows 22-70 (per-question
    detail). Both are tidied to long form (one row per theme[, question], benchmark).
    """
    wb = openpyxl.load_workbook(OVERALL_XLSX, read_only=True, data_only=True)
    ws = wb["Comparisons"]

    themes_df = _parse_comparisons_block(
        ws, header_row=11, data_start_row=12, data_end_row=19, has_question_col=False,
    )
    # The theme summary block repeats "Engagement" as both first and last row
    # (it's also the overall/aggregate theme) — drop the duplicate.
    themes_df = themes_df.drop_duplicates(subset=["theme", "benchmark"])

    questions_df = _parse_comparisons_block(
        ws, header_row=22, data_start_row=23, data_end_row=None, has_question_col=True,
    )
    wb.close()

    themes_df.to_csv(OUT / "theme_comparisons.csv", index=False)
    questions_df.to_csv(OUT / "question_comparisons.csv", index=False)
    print(f"theme_comparisons.csv: {len(themes_df)} rows")
    print(f"question_comparisons.csv: {len(questions_df)} rows")


def _extract_org_matrix(xlsx_path, granularity):
    """Shared parser for the Scores/Deltas org-unit matrices (Comp Div-Dept, Comp Sub-Div).

    Layout (both sheets): row 13 = org unit names, row 14 = response counts,
    then blocks of rows starting at 16. Each block opens with a header row
    ("Themes", or a theme name like "Purpose") whose data columns are all
    None, followed by data rows (theme scores, or question scores) whose
    first data column ("Overall") is numeric. Blocks are separated by blank
    rows. A header row's own label tells us whether we're in the theme
    summary block ("Themes") or a per-question block (the theme name).

    The Comp Div-Dept export pads its column range with the same org units
    repeated up to 3 times to fill out a fixed column count. Exactly one
    occurrence per unit carries the real response count and scores; the
    others are all-zero/'n/a' filler. A handful of units (below the survey's
    minimum-N reporting threshold) have n=0 in every occurrence — those are
    genuinely unreported, not padding artifacts. So for each unit we keep
    whichever occurrence has a non-zero response count, falling back to the
    first occurrence if every one of them is zero.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    scores_records = []
    deltas_records = []

    for sheet_name, records, value_col in [("Scores", scores_records, "score"),
                                             ("Deltas", deltas_records, "delta_pp")]:
        ws = wb[sheet_name]
        org_units_raw = _trim(next(_rows(ws, 13, 13)))[1:]  # skip leading None
        n_responses_raw = _trim(next(_rows(ws, 14, 14)))[1:]

        # For each unit name, pick the column with a non-zero response count
        # (the real data slot); fall back to the first column if all are zero.
        best_idx = {}
        for i, unit in enumerate(org_units_raw):
            if unit is None:
                continue
            n = n_responses_raw[i] if i < len(n_responses_raw) else None
            if unit not in best_idx:
                best_idx[unit] = i
            elif n not in (0, None) and n_responses_raw[best_idx[unit]] in (0, None):
                best_idx[unit] = i
        org_units = list(best_idx.keys())
        n_by_unit = {unit: n_responses_raw[i] for unit, i in best_idx.items()}
        first_idx = best_idx

        block_kind = None   # "theme" (Themes block) or "question" (a named theme's questions)
        current_theme = None
        for row in _rows(ws, 16):
            row = list(row)
            label = row[0]
            if label is None:
                continue

            is_header = all(v is None for v in row[1:len(org_units_raw) + 1])
            if is_header:
                if label == "Themes":
                    block_kind = "theme"
                    current_theme = None
                else:
                    block_kind = "question"
                    current_theme = label
                continue

            row_type = block_kind
            theme = label if block_kind == "theme" else current_theme
            question = None if block_kind == "theme" else label
            for unit, i in first_idx.items():
                val = row[i + 1] if i + 1 < len(row) else None
                if val is None:
                    continue
                records.append({
                    "granularity": granularity, "org_unit": unit,
                    "n_responses": n_by_unit.get(unit),
                    "row_type": row_type, "theme": theme, "question": question,
                    value_col: val,
                })

    wb.close()
    return pd.DataFrame(scores_records), pd.DataFrame(deltas_records)


def extract_org_matrices():
    scores_dd, deltas_dd = _extract_org_matrix(DIV_DEPT_XLSX, "division_department")
    scores_sd, deltas_sd = _extract_org_matrix(SUB_DIV_XLSX, "sub_division")

    scores = pd.concat([scores_dd, scores_sd], ignore_index=True)
    deltas = pd.concat([deltas_dd, deltas_sd], ignore_index=True)

    scores.to_csv(OUT / "org_scores.csv", index=False)
    deltas.to_csv(OUT / "org_deltas.csv", index=False)
    print(f"org_scores.csv: {len(scores)} rows")
    print(f"org_deltas.csv: {len(deltas)} rows")


def extract_comments():
    wb = openpyxl.load_workbook(COMMENTS_XLSX, read_only=True, data_only=True)
    ws = wb["Comments"]

    records = []
    for row in _rows(ws, 16):
        row = list(row)
        if not row or row[0] is None:
            continue
        question, engagement, cat1, cat2, cat3, sentiment, comment = (row + [None] * 7)[:7]
        records.append({
            "question": question,
            "engagement": engagement,
            "category_1": cat1 or None,
            "category_2": cat2 or None,
            "category_3": cat3 or None,
            "sentiment": sentiment,
            "comment": comment,
        })
    wb.close()

    df = pd.DataFrame(records)
    df.to_csv(OUT / "comments.csv", index=False)
    print(f"comments.csv: {len(df)} rows")
    return df


def derive_comment_themes(comments_df):
    """Join each comment to its implied survey theme(s) via sources/theme_category_map.csv.

    The map is a user-maintained editorial judgment call (which comment
    categories count as evidence for which of the 7 survey themes), not a
    derived or ID-based join — see the module docstring.
    """
    cat_map = pd.read_csv(THEME_CATEGORY_MAP_CSV)
    cat_to_theme = dict(zip(cat_map["category"], cat_map["theme"]))

    def themes_for(row):
        themes = []
        for col in ("category_1", "category_2", "category_3"):
            cat = row[col]
            theme = cat_to_theme.get(cat) if pd.notna(cat) else None
            if isinstance(theme, str) and theme and theme not in themes:
                themes.append(theme)
        return "|".join(themes)

    df = comments_df.copy()
    df["themes"] = df.apply(themes_for, axis=1)
    df.to_csv(OUT / "comments_themed.csv", index=False)
    print(f"comments_themed.csv: {len(df)} rows, {(df['themes'] != '').sum()} with a theme")


if __name__ == "__main__":
    extract_meta()
    extract_themes()
    extract_questions()
    extract_questions_5scale()
    extract_comparisons()
    extract_org_matrices()
    comments_df = extract_comments()
    derive_comment_themes(comments_df)
